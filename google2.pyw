# Import libraries, use pip install if you do not have them
import os
import docx
from PyPDF2 import PdfReader
import google.generativeai as genai
import tkinter as tk
from tkinter import scrolledtext
from tkinter import filedialog
import shutil
import requests
from bs4 import BeautifulSoup
import validators
import random
import time
import openpyxl
import traceback

# Function to get the API key from the environment variable you created
def get_api_key():
    """Retrieves the API key from the environment variables."""
    api_key = os.environ.get("GOOGLE_AI_API_KEY")
    if not api_key:
        raise ValueError("The GOOGLE_AI_API_KEY environment variable is not set.")
    return api_key

# Configure the API key for Google Generative AI (Gemini); keeping the print in the code is useful for debugging
try:
    genai.configure(api_key=get_api_key())
    print("API key configured successfully.")
except ValueError as e:
    print(f"Error configuring API key: {e}")
    exit()
except Exception as e:
    print(f"An unexpected error occurred during API key configuration: {e}")
    exit()

# Create a model instance
model = genai.GenerativeModel('gemini-2.0-flash') #look on google ai information for available models
print(f"Model initialized: {model}")

# Specify the directory path for document storage; this is mine, maybe you create your own path, as I do not use the virtual environment when running this one.
docs_dir = r'C:\Scripts\DOCS'

def read_document_contents(directory):
    """Reads the contents of all supported documents in the given directory."""
    document_contents = []
    for filename in os.listdir(directory):
        filepath = os.path.join(directory, filename)

        if filename.endswith(".txt"):  # For .txt files
            try:
                with open(filepath, 'r', encoding='utf-8') as file:
                    content = file.read()
                    document_contents.append(content)
            except Exception as e:
                print(f"Error reading {filename}: {e}")

        elif filename.endswith(".docx"):  # For .docx files
            try:
                doc = docx.Document(filepath)
                content = "\n".join([paragraph.text for paragraph in doc.paragraphs])
                document_contents.append(content)
            except Exception as e:
                print(f"Error reading {filename}: {e}")

        elif filename.endswith(".pdf"):  # For .pdf files
            try:
                pdf_reader = PdfReader(filepath)
                content = "\n".join([page.extract_text() for page in pdf_reader.pages])
                document_contents.append(content)
            except Exception as e:
                print(f"Error reading {filename}: {e}")

        elif filename.endswith(".xlsx"):  # For .xlsx files (Excel)
            try:
                workbook = openpyxl.load_workbook(filepath)
                content = ""
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows():
                        cell_values = [str(cell.value) for cell in row if cell.value is not None]
                        content += ", ".join(cell_values) + "\n"
                document_contents.append(content)
            except Exception as e:
                print(f"Error reading {filename}: {e}")

        else:
            print(f"Unsupported file type: {filename}")

    return document_contents

def fetch_and_extract_text_from_url(url):
    """Fetches content from a URL and extracts text using BeautifulSoup."""
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()

        soup = BeautifulSoup(response.content, 'html.parser')
        text_content = soup.get_text(separator="\n")
        return text_content.strip()
    except requests.exceptions.RequestException as e:
        return f"Error fetching URL: {e}"
    except Exception as e:
        return f"Error processing URL: {e}"

def is_relevant_document(query, document_contents):
    """
    Checks if any document in `document_contents` is relevant to the query.
    Uses basic keyword matching (can be improved with more sophisticated NLP techniques).
    """
    query_keywords = set(query.lower().split())
    for content in document_contents:
        content_keywords = set(content.lower().split())
        if query_keywords & content_keywords:  # Intersection of keywords
            return True
    return False

def generate_dynamic_prompt(query, document_contents, url_content=None, previous_context=None):
    """
    Dynamically generates a prompt based on the relevance of documents and context.
    """
    relevant_documents = [doc for doc in document_contents if is_relevant_document(query, [doc])]

    if relevant_documents:
        combined_content = "\n\n".join(relevant_documents)
        prompt = f"Use the following document content to answer the question:\n\n{combined_content}\n\n"
    else:
        prompt = "Use your general knowledge or search the web to answer the following question.\n"
        if url_content:
            prompt += f"Consider this URL content if relevant:\n{url_content}\n\n"
    
    if previous_context:
        prompt += f"Previous context:\n{previous_context}\n\n"
    
    prompt += f"Question: {query}\n"
    prompt += "Provide a concise, accurate, and relevant answer."

    return prompt

def generate_answer(query, document_contents, url_content=None, previous_context=None):
    """
    Generates an answer dynamically based on relevance of documents and fallback mechanisms.
    """
    dynamic_prompt = generate_dynamic_prompt(query, document_contents, url_content, previous_context)

    print("DEBUG: Full dynamic prompt being sent to the AI:")
    print(dynamic_prompt)

    try:
        response = model.generate_content(dynamic_prompt)
        print(f"DEBUG: Response received. Type: {type(response)}")
        return response.text.strip()
    except Exception as e:
        print(f"An error occurred in generate_answer: {e}")
        return "Failed to generate answer."

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Dynamic AI Assistant")
        
        self.document_contents = read_document_contents(docs_dir)
        self.chat_history = []

        self.text_area = scrolledtext.ScrolledText(root, wrap=tk.WORD, width=80, height=20)
        self.text_area.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
        self.text_area.config(state=tk.DISABLED)

        self.query_label = tk.Label(root, text="What would you like to ask (or provide a URL)?")
        self.query_label.pack()
        
        self.query_entry = tk.Entry(root, width=50)
        self.query_entry.pack(pady=5)
        
        self.query_entry.bind('<Return>', self.ask_question)

        self.upload_button = tk.Button(root, text="Upload Document", command=self.upload_document)
        self.upload_button.pack(pady=5)

        self.exit_button = tk.Button(root, text="EXIT", command=self.root.destroy)
        self.exit_button.pack(pady=5)

        self.restart_button = tk.Button(root, text="RESTART", command=self.restart)
        self.restart_button.pack(pady=5)

        self.export_button = tk.Button(root, text="Export Chat Log", command=self.export_chat_log)
        self.export_button.pack(pady=5)

    def ask_question(self, event=None):
        query = self.query_entry.get().strip()
        
        if validators.url(query):
            self.display_text(f"Fetching content from URL: {query}\n")
            url_content = fetch_and_extract_text_from_url(query)
            if "Error" in url_content:
                self.display_text(url_content + "\n")
                return
        else:
            url_content = None

        self.display_text(f"Question: {query}\n")
        
        answer = generate_answer(
            query,
            self.document_contents,
            url_content,
            previous_context="\n".join([f"{msg['question']}\n{msg['answer']}" for msg in self.chat_history[-5:]])
        )

        self.display_text(f"Answer: {answer}\n\n{'-'*50}\n")
        
        self.query_entry.delete(0, tk.END)
        
        self.chat_history.append({"question": query, "answer": answer})

    def display_text(self, text):
        """Displays text in the UI."""
        self.text_area.config(state=tk.NORMAL)
        self.text_area.insert(tk.END, text)
        self.text_area.config(state=tk.DISABLED)
        self.text_area.see(tk.END)

    def upload_document(self):
        filepath = filedialog.askopenfilename(title="Select Document to Upload")
        
        if filepath:
            try:
                shutil.copy(filepath, docs_dir)
                self.display_text(f"Uploaded {os.path.basename(filepath)} to {docs_dir}\n")
                self.document_contents = read_document_contents(docs_dir)  # Refresh documents after upload
            except Exception as e:
                self.display_text(f"Error uploading document: {e}\n")

    def restart(self):
        """Restarts the application."""
        self.root.destroy()
        
        root = tk.Tk()
        
        app = App(root)
        
        root.mainloop()

    def export_chat_log(self):
        """Exports chat log to a file."""
        
        filename = filedialog.asksaveasfilename(defaultextension=".txt",
                                                filetypes=[("Text Files", "*.txt"),
                                                           ("All Files", "*.*")])
        
        if filename:
            try:
                # Get chat log from UI
                chat_log = self.text_area.get("1.0", tk.END).strip()
                
                with open(filename, "w", encoding="utf-8") as f:
                    f.write(chat_log)
                
                self.display_text(f"Chat log saved to {filename}\n")
            
            except Exception as e:
                self.display_text(f"Error saving chat log: {e}\n")

def main():
    root = tk.Tk()
    
    app = App(root)
    
    root.mainloop()

if __name__ == "__main__":
    main()
